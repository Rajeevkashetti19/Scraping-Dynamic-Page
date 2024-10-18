import streamlit as st
import pandas as pd
from typing import List, Dict, Tuple
from collections import defaultdict
import time
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import random

class Employee:
    def __init__(self, id: str, grade: str, manager: str, business: str, city: str):
        self.id = id
        self.grade = grade
        self.manager = manager
        self.business = business
        self.city = city
        self.team = set()

def build_team_hierarchy(employees: List[Employee]) -> None:
    emp_dict = {emp.id: emp for emp in employees}
    for emp in employees:
        current = emp
        while current.manager and current.manager in emp_dict:
            emp_dict[current.manager].team.add(emp.id)
            current = emp_dict[current.manager]

def group_by_business_and_grade(df: pd.DataFrame, min_group: int, max_group: int) -> List[List[str]]:
    groups = []
    for business in df['Business'].unique():
        business_df = df[df['Business'] == business]
        for grade in business_df['Grade'].unique():
            grade_df = business_df[business_df['Grade'] == grade]
            employee_ids = grade_df['EmployeeID'].tolist()
            while len(employee_ids) >= min_group:
                group = employee_ids[:max_group]
                groups.append(group)
                employee_ids = employee_ids[max_group:]
    return groups

def group_by_city_and_grade(df: pd.DataFrame, num_months: int) -> List[List[str]]:
    groups = []
    for city in df['City'].unique():
        city_df = df[df['City'] == city]
        for grade in city_df['Grade'].unique():
            grade_df = city_df[city_df['Grade'] == grade]
            employee_ids = grade_df['EmployeeID'].tolist()
            num_groups = max(1, len(employee_ids) // num_months)
            for i in range(num_groups):
                start = i * num_months
                end = start + num_months
                group = employee_ids[start:end]
                if group:
                    groups.append(group)
    return groups

def group_employees_complex(df: pd.DataFrame) -> Tuple[List[List[str]], Dict[str, str]]:
    # ... [The complex grouping function remains the same] ...

def create_download_link(df, filename, color_coded=False):
    output = io.BytesIO()
    
    if color_coded:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
        
        # Write headers
        for col, header in enumerate(df.columns, start=1):
            sheet.cell(row=1, column=col, value=header)
        
        # Write data and apply color coding
        colors = [PatternFill(start_color=f"00{random.randint(0, 0xFFFFFF):06x}", end_color=f"00{random.randint(0, 0xFFFFFF):06x}", fill_type="solid") for _ in range(len(df))]
        
        for row, (_, data_row) in enumerate(df.iterrows(), start=2):
            for col, value in enumerate(data_row, start=1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.fill = colors[row-2]
        
        workbook.save(output)
    else:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
    
    excel_data = output.getvalue()
    return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{pd.compat.base64.b64encode(excel_data).decode()}" download="{filename}">Download {filename}</a>'

def main():
    st.title("Employee Grouping Application")

    uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")
    
    if uploaded_file is not None:
        progress_text = st.empty()
        progress_bar = st.progress(0)

        progress_text.text("Loading data...")
        progress_bar.progress(10)
        time.sleep(0.5)

        df = pd.read_excel(uploaded_file)
        
        progress_text.text("Data loaded successfully!")
        progress_bar.progress(30)
        time.sleep(0.5)

        st.write("Data Preview:")
        st.dataframe(df.head())

        grouping_options = st.multiselect(
            "Select grouping method(s)",
            ["Group by Business and Grade", "Group by City and Grade", "Complex Grouping (C15/C16 leaders)"],
            default=["Group by Business and Grade"]
        )

        # Conditional inputs based on selected grouping methods
        if "Group by Business and Grade" in grouping_options:
            min_group = st.number_input("Minimum group size (for Business and Grade grouping)", min_value=2, max_value=10, value=3)
            max_group = st.number_input("Maximum group size (for Business and Grade grouping)", min_value=min_group, max_value=10, value=4)
        
        if "Group by City and Grade" in grouping_options:
            num_months = st.number_input("Number of months for grouping (for City and Grade grouping)", min_value=1, max_value=12, value=3)

        if st.button("Generate Groups"):
            for option in grouping_options:
                progress_text.text(f"Generating groups using {option}...")
                progress_bar.progress(50)
                time.sleep(0.5)

                if option == "Group by Business and Grade":
                    progress_text.text("Grouping employees by Business and Grade...")
                    groups = group_by_business_and_grade(df, min_group, max_group)
                    grouped_df = pd.DataFrame([(group[0], ', '.join(group)) for group in groups], columns=['Group Leader', 'Group Members'])
                    st.subheader("Groups by Business and Grade")
                    st.dataframe(grouped_df)
                    st.markdown(create_download_link(grouped_df, "business_grade_groups.xlsx"), unsafe_allow_html=True)
                
                elif option == "Group by City and Grade":
                    progress_text.text("Grouping employees by City and Grade...")
                    groups = group_by_city_and_grade(df, num_months)
                    grouped_df = pd.DataFrame([(group[0], ', '.join(group)) for group in groups], columns=['Group Leader', 'Group Members'])
                    st.subheader("Groups by City and Grade")
                    st.dataframe(grouped_df)
                    st.markdown(create_download_link(grouped_df, "city_grade_groups.xlsx", color_coded=True), unsafe_allow_html=True)
                
                elif option == "Complex Grouping (C15/C16 leaders)":
                    progress_text.text("Performing complex grouping with C15/C16 leaders...")
                    groups, ungrouped_reasons = group_employees_complex(df)
                    grouped_df = pd.DataFrame([(group[0], ', '.join(group)) for group in groups], columns=['Group Leader', 'Group Members'])
                    st.subheader("Complex Grouping (C15/C16 leaders)")
                    st.dataframe(grouped_df)
                    st.markdown(create_download_link(grouped_df, "complex_groups.xlsx"), unsafe_allow_html=True)
                    
                    st.subheader("Ungrouped C15/C16 employees and reasons:")
                    for emp_id, reason in ungrouped_reasons.items():
                        emp = df[df['EmployeeID'] == emp_id].iloc[0]
                        st.text_area(f"Employee ID: {emp_id}", 
                                     f"Grade: {emp['Grade']}, Business: {emp['Business']}, City: {emp['City']}\nReason: {reason}",
                                     height=100)

                progress_bar.progress(90)
                time.sleep(0.5)

            progress_text.text("Grouping complete!")
            progress_bar.progress(100)

if __name__ == "__main__":
    main()